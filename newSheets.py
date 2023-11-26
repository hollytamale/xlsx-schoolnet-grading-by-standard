import openpyxl as xl
import schoolnetFunctions
from schoolnetFunctions import pull_test_id, pull_names_from_schoolnet, pull_point_columns, q_resubmit_standards
from classes import Student

wb = xl.load_workbook('2023-24Grades.xlsx')
# sheet = wb.active

file_to_import = 'TestResults_5174646.xlsx'   # Replace additional file name here
testid = pull_test_id(file_to_import)
local_standard_list = []
Sheet1 = wb["Sheet1"]

if wb.sheetnames.count(testid):
    sheet_new = wb[testid]
else:
    sheet_new = wb.create_sheet(testid, 1)

# names_array = pull_names_from_schoolnet(file_to_import)
names_array = pull_names_from_schoolnet(file_to_import)
points_array = pull_point_columns(file_to_import)

schoolnetFunctions.unpack_names_from_schoolnet(names_array, sheet_new)
schoolnetFunctions.unpack_point_columns(points_array, sheet_new, names_array)

q_resubmit_standards(sheet_new, points_array[1], names_array[1], local_standard_list, testid)

# Last thing is to update the main sheet with new "objects" if needed
schoolnetFunctions.unpack_names_from_schoolnet(names_array, Sheet1)

"""Notes to future noodle, 11/18:
Generalize pullnames function and use it to compare names that exist in unpacknames function.
trying to create list of names that doesn't double, and adds new names to mainsheet. I'm nto sure
otherwise how to compare names taht are already in the spreadsheet. I don't think I can search rows like
I can search list items. Not sure though... I mean, it's alphabetical. I could search and add rows.
Not sure wjhich is easier.
Add to list of names if not there.


11/6:
- still struggling to make standards list universal
    - add standards to list, don't double. Don't double between new sheets, either
        - this is easy -- if it exists already, don't add it. Still... universal? Where is the list initiated? Not sure.

    - how to associate with kids? and with standards? and to recalculate with new quizzes?
        - Do I associate the quiz ID with each data point for each kid?
        - *** each kid object has the group of standards in an array. Each standard is associated with their answers.
            - each standard also averages their available answers. 
            - Each question is also associated with the test number, for cross-referencing purposes.
        - to calculate when new quizzes are added... fine. But when a quiz is resubmitted? Should it delete their old
            answers, or somehow just add answers to arrays for kids who hadn't taken the quiz? 2nd is better... How?
            - Maybe quizzes are also objects, and list of kids are associated with these objects.
            
- wondering how to populate main sheet
- wanting to clean up "q_resubmit_standards" bc it seems there are several useful bits that could be used elsewhere.
    For examples, 
    - question col pos might be useful (without offsetting automatically by 3 cols)
    - could use column offsets function within this? Universal offset function? 
"""

wb.save('2023-24Grades.xlsx')