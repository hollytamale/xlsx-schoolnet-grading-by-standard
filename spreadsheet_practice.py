import openpyxl as xl
from schoolnetFunctions import pull_test_id, pull_names_from_schoolnet, pull_point_columns, standards_no_repeat

"""NOTES TO THE NOODLE:
10/18
- Same sheet, imported again, messes up. How to note completed? Symbol? "DONE" ? 
- Also, Maybe change color of averages columns
- What about when a new sheet is imported... how does this work with multiple sheets?

- Next steps, 
    - standardizing Standards inputs
    - Consolidating student info into one tag? Good idea, I think?
    - Consolidate info by average or standard on first sheet, sheet 1.
    - lastly, major clean up and reorganizing. Hm. 
        - use classes instead and share variables
        - Person class? With standards attributes
"""

# wb = xl.load_workbook('2023-24Grades.xlsx')
# sheet = wb.active
#
# file_to_import = 'TestResults_5174646.xlsx'   # Replace additional file name here
# testid = pull_test_id(file_to_import)

# if wb.sheetnames.count(testid) == True:
#     sheet_new = wb[testid]
# else:
#     sheet_new = wb.create_sheet(testid, 1)

# # Unpack array of redacted names and IDs to a new sheet from pull_names function
# names_array = pull_names_from_schoolnet(file_to_import)
# for item in range(0, len(names_array)):
#     names_row = names_array[item]
#     for point in range(len(names_row)):
#         sheet_new.cell(item + 1, point + 1).value = str(names_array[item][point])

# # Unpack points array to same test sheet from pull_points function
# points_array = pull_point_columns(file_to_import)
# for item in range(0, len(points_array)):
#     points_row = points_array[item]
#     for point in range(0, len(points_row)):
#         sheet_new.cell(item + 1, point + len(names_row) + 1).value = points_array[item][point]



# standard_list = []
# # Need to standardize how standards are input, or they won't categorize together
# if sheet_new.cell(1, 1).value:
#     resubmit = input("Do you wish to resubmit standards for this quiz? (Yes/No): ").lower()
#     if resubmit == "yes":
#         # This junk is redundant. I need to avoid repeating myself.. laterz.
#         sheet_new.insert_rows(1)
#         print(f"For test ID #{testid}, identify standard for: ")
#         for question in range(len(points_row)):
#             q_standard = input(f"Q{question + 1}: ")
#             ques_num = "Q" + str(question + 1)  # Formats question number
#             ques_column_pos = question + len(names_row) + 1
#             sheet_new.cell(1, ques_column_pos).value = f'{q_standard}'
#             standards_no_repeat(q_standard, standard_list)      # Modded with function, but not sure about next one...
#     else:
#         print("Okay, no changes will be made.")

# global_standard_list = list()    # Can I move elsewhere to not reset this every time?
# standards_no_repeat(standard_list, global_standard_list)

# Inserting new columns
col_count = 0
column_offset = len(names_row) + col_count + 1
for standard in standard_list:
    sheet_new.insert_cols(column_offset)
    col_count += 1

standard_pos = []
single_standard_pos = []
# col_count = 0

# Finding col nums
for standard in standard_list:
    single_standard_pos = []
    for header_cols in range(1, sheet_new.max_column + 1):
        if sheet_new.cell(1, header_cols).value == standard:
            single_standard_pos.append(header_cols)
    standard_pos.append(single_standard_pos)

# Add names, oh no
# Adding col one pos too far right
col_count = 0
for standard in standard_list:
    sheet_new.cell(1, column_offset + col_count).value = standard
    col_count += 1

col_count = 0
for name in range(1, sheet_new.max_row + 1):    # should this start at 2?
    off_offset = 0
    for cols in standard_pos:   # pulls each standard's column(s)
        occurs_count = 0
        point_sum = 0
        sum_per_standard = 0
        standard_avg = 0
        for i in cols:          # pulls values for each standard
            try:
                sum_per_standard += sheet_new.cell(name + 1, i).value
                occurs_count += 1
            except:
                name += 1
        try:
            standard_avg = sum_per_standard / occurs_count
        except ZeroDivisionError:
            standard_avg = 0
            print("Division by zero")
        sheet_new.cell(name + 1, column_offset + off_offset).value = standard_avg
        off_offset += 1




""" On first sheet:
- need list of all names
- match names to scores... Skip those who don't have a score.
- average scores from each sheet as added -- active process! How? 
    - As new quizzes are added, add them all up on each sheet, don't divide? 
    - Keep track of how many col_names there are for all sheets?
    - Divide them out only on main sheet?
- Once a kid takes a quiz, I should be able to upload the same spreadsheet again without errors for others.
"""

# Working in the main sheet
col = 0
main_sheet_headers = ['Name']
for i in main_sheet_headers:
    col += 1        # This can be... simpler, right?
    sheet.cell(1, col).value = i

# only append standards to standard list if not already in existence
# main_sheet_headers.append(standard_list)





# wb.save('2023-24Grades.xlsx')
