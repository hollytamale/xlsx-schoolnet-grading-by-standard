import openpyxl as xl
# from schoolnetFunctions import pull_test_id, pull_names_from_schoolnet, pull_point_columns, standards_no_repeat
# from classes import Student

# wb = xl.load_workbook('2023-24Grades.xlsx')
# sheet = wb.active

# file_to_import = 'TestResults_5174646.xlsx'   # Replace additional file name here
# testid = pull_test_id(file_to_import)

"""
- Maybe remove bits that add data back to spreadsheet?
"""
def pull_names_from_schoolnet(new_schoolnet_file):
    wb = xl.load_workbook(new_schoolnet_file)
    sheet = wb.active
    # Store in array/list
    names_array = []
    for row in range(1, sheet.max_row):
        data_list = []
        for cell in range(1, 2):
            last_name = sheet.cell(row + 1, 1).value
            first_name = sheet.cell(row + 1, 2).value
            id_number = str(sheet.cell(row + 1, 3).value)
            data_list = [last_name, first_name, id_number]
            names_array.append(data_list)
    return names_array


# Fix so that it unpacks names that don't exist. Compare it to list from "pull names" function, used a 2nd time.
def unpack_names_from_schoolnet(names_array, sheet):
    names_array.sort()
    for item in range(0, len(names_array)):
        if sheet.cell(row=item+1, column=3).value == names_array[item][2]:
            pass
        else:
            names_row = names_array[item]
            for point in range(len(names_row)):
                sheet.cell(item + 2, point + 1).value = str(names_array[item][point])


def pull_point_columns(new_schoolnet_file):
    wb = xl.load_workbook(new_schoolnet_file)
    sheet = wb.active
    # Generate list of N + # (less than number of columns)
    n_column = ["P" + str(num) for num in range(sheet.max_column + 1)]
    point_column_list = []
    # Find start and end of P columns for range, rather than set values
    for column in range(1, sheet.max_column + 1):
        col_header = sheet.cell(1, column)
        for value in n_column:
            if value == col_header.value:
                point_column_list.append(column)
                point_column_min = point_column_list[0]
                point_column_max = point_column_list[-1]
    # Found columns for points -- example, min is column 23 and max is column 30.
    # Now, find and store values for all rows from those columns.
    points_array = []
    for row in range(2, sheet.max_row + 1):
        points_list = []
        for i in range((point_column_max + 1) - point_column_min):
            point_column = point_column_min + i
            point_thing = sheet.cell(row, point_column).value
            points_list += [point_thing]
        points_array.append(points_list)
    return points_array


def unpack_point_columns(points_array, sheet, names_column):
    for item in range(0, len(points_array)):
        points_row = points_array[item]
        for point in range(0, len(points_row)):
            sheet.cell(item + 2, point + len(names_column[1]) + 1).value = points_array[item][point]

# def unpack_items_from_schoolnet(names_array, points_array, sheet):
#     for item in range(0, len)


def pull_test_id(new_schoolnet_file):
    wb = xl.load_workbook(new_schoolnet_file)
    sheet = wb.active
    testid = str(sheet.cell(2, 8).value)
    return testid


# def standards_no_repeat(standards_to_input, standard_list):
#     for standard in standards_to_input:
#         if standard_list.count(standard) < 1:
#             standard_list.append(standard)


# Need to standardize how standards are input, or they won't categorize together
def q_resubmit_standards(sheet, points_row, names_row, local_standard_list, testid):
    if sheet.cell(1, 1).value == str("*"):
        resubmit = input("This quiz already has standards associated. Would you like to resubmit? (Yes/No): ").lower()
    else:
        sheet.insert_rows(1)
        resubmit = input("Do you wish to submit standards for this quiz? (Yes/No): ").lower()
    if resubmit == "yes":
        sheet.cell(1, 1).value = str("*")
        print(f"For test ID #{testid}, identify standard for: ")
        for question in range(len(points_row)):
            q_standard = input(f"Q{question + 1}: ")
            # ques_num = "Q" + str(question + 1)  # Formats question number
            ques_column_pos = question + len(names_row) + 1
            sheet.cell(1, ques_column_pos).value = f'{q_standard}'
            if local_standard_list.count(q_standard) < 1:
                local_standard_list.append(q_standard)
    else:
        print("Okay, no changes will be made.")
